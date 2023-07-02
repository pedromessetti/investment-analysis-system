# scrapers/invest_site.py

import openpyxl
import os
import csv
import time
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager

def download_invest_site_xlsx():

    # Appropriated URL of Invest Site to scrape
    url = 'https://www.investsite.com.br/selecao_acoes.php?dt_arr=%255B%252220230623%2522%252C%2522atual%2522%255D&todos=todos&ROTanC_min=&ROTanC_max=&chk_lst%5B%5D=itm7&ROInvC_min=&ROInvC_max=&chk_lst%5B%5D=itm8&ROE_min=&ROE_max=&chk_lst%5B%5D=itm9&ROA_min=&ROA_max=&chk_lst%5B%5D=itm10&margem_liquida_min=&margem_liquida_max=&chk_lst%5B%5D=itm11&margem_bruta_min=&margem_bruta_max=&chk_lst%5B%5D=itm12&margem_EBIT_min=&margem_EBIT_max=&chk_lst%5B%5D=itm13&giro_ativo_min=&giro_ativo_max=&chk_lst%5B%5D=itm14&fin_leverage_min=&fin_leverage_max=&chk_lst%5B%5D=itm15&debt_equity_min=&debt_equity_max=&chk_lst%5B%5D=itm16&p_e_min=&p_e_max=&chk_lst%5B%5D=itm17&p_bv_min=&p_bv_max=&chk_lst%5B%5D=itm18&p_receita_liquida_min=&p_receita_liquida_max=&chk_lst%5B%5D=itm19&p_FCO_min=&p_FCO_max=&chk_lst%5B%5D=itm20&p_FCF1_min=&p_FCF1_max=&chk_lst%5B%5D=itm21&p_EBIT_min=&p_EBIT_max=&chk_lst%5B%5D=itm22&p_ncav_min=&p_ncav_max=&chk_lst%5B%5D=itm23&p_ativo_total_min=&p_ativo_total_max=&chk_lst%5B%5D=itm24&p_capital_giro_min=&p_capital_giro_max=&chk_lst%5B%5D=itm25&EV_EBIT_min=&EV_EBIT_max=&chk_lst%5B%5D=itm26&EV_EBITDA_min=&EV_EBITDA_max=&chk_lst%5B%5D=itm27&EV_receita_liquida_min=&EV_receita_liquida_max=&chk_lst%5B%5D=itm28&EV_FCO_min=&EV_FCO_max=&chk_lst%5B%5D=itm29&EV_FCF1_min=&EV_FCF1_max=&chk_lst%5B%5D=itm30&EV_ativo_total_min=&EV_ativo_total_max=&chk_lst%5B%5D=itm31&div_yield_min=&div_yield_max=&chk_lst%5B%5D=itm32&vol_financ_min=&vol_financ_max=&chk_lst%5B%5D=itm33&market_cap_min=&market_cap_max=&chk_lst%5B%5D=itm34&setor='

    driver, wait = setup_firefox_driver()

    driver.get(url)
    get_url = driver.current_url
    wait.until(EC.url_to_be(url))

    if get_url == url:
        button = wait.until(EC.visibility_of_element_located((By.ID, 'botao_arquivo')))
    button.click()

    downloads_folder = os.path.expanduser('~/Downloads/')
    # Wait for the file to be downloaded
    while True:
        files = os.listdir(downloads_folder)
        xlsx_files = [f for f in files if f.endswith('.xlsx') and f.startswith('Stock_Screener_')]
        if len(xlsx_files) > 0:
            break
        time.sleep(5)

    # Find the name of the downloaded XLSX file in the Downloads folder
    file_name = next(filename for filename in os.listdir(downloads_folder) if filename.endswith('.xlsx') and filename.startswith('Stock_Screener_'))

    # Move the file from Downloads folder to the current folder
    shutil.move(os.path.join(downloads_folder, file_name), 'csv/invest_site.xlsx')
    print('invest_site.xlsx successfully downloaded!')
    
    driver.quit()

    convert_xlsx_to_csv()


def setup_firefox_driver():
    try:
        driver = webdriver.Firefox(service=Service(executable_path=GeckoDriverManager().install()))
        wait = WebDriverWait(driver, 10)
        return driver, wait
    except:
        print('Erro ao iniciar o driver do Firefox')
        exit()



def convert_xlsx_to_csv():

    # Open the XLSX file
    workbook = openpyxl.load_workbook('csv/invest_site.xlsx')

    # Select the first worksheet
    worksheet = workbook.active

    # Delete the first 2 rows
    worksheet.delete_rows(1, 2)
    
    # Delete the last 38 rows
    worksheet.delete_rows(worksheet.max_row - 37, worksheet.max_row)

    # Open the CSV file for writing
    with open('csv/invest_site.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)

        # Loop through all rows in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == "NA":
                    cell.value = 0
            # Extract the values from each cell in the row
            values = [cell.value for cell in row]

            # Write the values to the CSV file
            writer.writerow(values)
    
    print('invest_site.csv successfully converted!')
